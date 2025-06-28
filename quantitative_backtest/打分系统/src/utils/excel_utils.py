# -*- coding: utf-8 -*-
"""
Excel输出工具模块
负责将股票评分结果格式化输出为Excel文件

项目：基本面量化回测系统 - 打分系统
作者：量化团队
创建时间：2025-06-28
"""

import pandas as pd
import numpy as np
from datetime import datetime
from typing import Dict, List, Any, Optional
from pathlib import Path
import logging
import warnings
warnings.filterwarnings('ignore')

# 导入配置
try:
    from config.scoring_config import (
        get_output_config, get_columns_config, 
        get_basic_config, get_logging_config
    )
except ImportError as e:
    print(f"警告：无法导入配置: {e}")

# 检查并导入Excel相关库
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.formatting.rule import CellIsRule
    EXCEL_AVAILABLE = True
except ImportError:
    print("警告：openpyxl未安装，Excel功能将受限")
    EXCEL_AVAILABLE = False

# 设置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ==============================================================================
# 1. Excel格式化器
# ==============================================================================

class ExcelFormatter:
    """Excel格式化器"""
    
    def __init__(self):
        self.output_config = get_output_config()
        self.columns_config = get_columns_config()
        self.basic_config = get_basic_config()
        
        # 样式定义
        self.styles = self._create_styles()
    
    def _create_styles(self) -> Dict[str, Any]:
        """创建Excel样式"""
        if not EXCEL_AVAILABLE:
            return {}
        
        styles = {
            # 标题样式
            'header': {
                'font': Font(bold=True, size=12, color='FFFFFF'),
                'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                ),
                'alignment': Alignment(horizontal='center', vertical='center')
            },
            
            # 数据样式
            'data': {
                'font': Font(size=10),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                ),
                'alignment': Alignment(horizontal='center', vertical='center')
            },
            
            # 排名样式（前10名高亮）
            'top_rank': {
                'font': Font(bold=True, size=10, color='FFFFFF'),
                'fill': PatternFill(start_color='00B050', end_color='00B050', fill_type='solid'),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                ),
                'alignment': Alignment(horizontal='center', vertical='center')
            },
            
            # 股票代码样式
            'stock_code': {
                'font': Font(bold=True, size=10),
                'alignment': Alignment(horizontal='center', vertical='center')
            },
            
            # 数值样式
            'number': {
                'font': Font(size=10),
                'alignment': Alignment(horizontal='right', vertical='center'),
                'number_format': '0.0000'
            },
            
            # 百分比样式
            'percentage': {
                'font': Font(size=10),
                'alignment': Alignment(horizontal='right', vertical='center'),
                'number_format': '0.00%'
            }
        }
        
        return styles
    
    def format_dataframe_for_excel(self, df: pd.DataFrame) -> pd.DataFrame:
        """格式化DataFrame用于Excel输出"""
        if df.empty:
            return df
        
        formatted_df = df.copy()
        
        # 数值列格式化
        numeric_columns = formatted_df.select_dtypes(include=[np.number]).columns
        for col in numeric_columns:
            if col not in ['final_rank']:
                formatted_df[col] = pd.to_numeric(formatted_df[col], errors='coerce')
                formatted_df[col] = formatted_df[col].round(4)
        
        # 确保排名列为整数
        if 'final_rank' in formatted_df.columns:
            formatted_df['final_rank'] = formatted_df['final_rank'].fillna(0).astype(int)
        
        # 字符串列去除多余空格
        string_columns = formatted_df.select_dtypes(include=['object']).columns
        for col in string_columns:
            if col in formatted_df.columns:
                formatted_df[col] = formatted_df[col].astype(str).str.strip()
        
        return formatted_df
    
    def get_column_order(self) -> List[str]:
        """获取Excel输出的列顺序"""
        all_columns = []
        
        # 添加各类列
        for column_group in ['basic_info', 'profitability_factors', 'growth_factors', 
                           'safety_factors', 'dimension_scores', 'final_results']:
            if column_group in self.columns_config:
                columns = [item['column'] for item in self.columns_config[column_group]]
                all_columns.extend(columns)
        
        return all_columns
    
    def get_column_names_mapping(self) -> Dict[str, str]:
        """获取列名映射（英文->中文）"""
        name_mapping = {}
        
        for column_group in self.columns_config.values():
            for item in column_group:
                name_mapping[item['column']] = item['name']
        
        return name_mapping
    
    def get_column_widths(self) -> Dict[str, int]:
        """获取列宽设置"""
        width_mapping = {}
        
        for column_group in self.columns_config.values():
            for item in column_group:
                width_mapping[item['column']] = item.get('width', 12)
        
        return width_mapping

# ==============================================================================
# 2. Excel写入器
# ==============================================================================

class ExcelWriter:
    """Excel写入器"""
    
    def __init__(self):
        self.formatter = ExcelFormatter()
        self.output_config = get_output_config()
    
    def write_scoring_results(self, results: pd.DataFrame, 
                            calculation_date: str,
                            statistics: Dict[str, Any],
                            output_path: Optional[str] = None) -> str:
        """
        写入评分结果到Excel文件
        
        Args:
            results: 评分结果DataFrame
            calculation_date: 计算日期
            statistics: 统计信息
            output_path: 输出路径，如果为None则自动生成
            
        Returns:
            str: 输出文件路径
        """
        if results.empty:
            raise ValueError("评分结果为空，无法生成Excel文件")
        
        # 生成输出路径
        if output_path is None:
            output_path = self._generate_output_path(calculation_date)
        
        logger.info(f"开始写入Excel文件: {output_path}")
        
        try:
            if EXCEL_AVAILABLE:
                self._write_with_formatting(results, calculation_date, statistics, output_path)
            else:
                self._write_simple_excel(results, output_path)
            
            logger.info(f"✅ Excel文件写入成功: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"❌ Excel文件写入失败: {e}")
            raise
    
    def _generate_output_path(self, calculation_date: str) -> str:
        """生成输出文件路径"""
        # 确保输出目录存在
        output_dir = Path(self.output_config['output_directory'])
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # 生成文件名
        date_str = calculation_date.replace('-', '')
        filename = self.output_config['filename_template'].format(date=date_str)
        
        return str(output_dir / filename)
    
    def _write_with_formatting(self, results: pd.DataFrame, 
                             calculation_date: str, 
                             statistics: Dict[str, Any], 
                             output_path: str):
        """写入带格式的Excel文件"""
        # 格式化数据
        formatted_results = self.formatter.format_dataframe_for_excel(results)
        
        # 重新排列列顺序
        column_order = self.formatter.get_column_order()
        available_columns = [col for col in column_order if col in formatted_results.columns]
        formatted_results = formatted_results[available_columns]
        
        # 重命名列
        column_names = self.formatter.get_column_names_mapping()
        formatted_results = formatted_results.rename(columns=column_names)
        
        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.output_config['excel_settings']['sheet_name']
        
        # 写入数据
        for r in dataframe_to_rows(formatted_results, index=False, header=True):
            ws.append(r)
        
        # 应用格式
        self._apply_excel_formatting(ws, len(formatted_results), statistics)
        
        # 保存文件
        wb.save(output_path)
    
    def _write_simple_excel(self, results: pd.DataFrame, output_path: str):
        """写入简单Excel文件（不带格式）"""
        # 格式化数据
        formatted_results = self.formatter.format_dataframe_for_excel(results)
        
        # 重新排列列顺序和重命名
        column_order = self.formatter.get_column_order()
        available_columns = [col for col in column_order if col in formatted_results.columns]
        formatted_results = formatted_results[available_columns]
        
        column_names = self.formatter.get_column_names_mapping()
        formatted_results = formatted_results.rename(columns=column_names)
        
        # 保存为Excel
        formatted_results.to_excel(output_path, index=False, sheet_name='股票评分结果')
    
    def _apply_excel_formatting(self, worksheet, data_rows: int, statistics: Dict[str, Any]):
        """应用Excel格式"""
        if not EXCEL_AVAILABLE:
            return
        
        # 设置列宽
        column_widths = self.formatter.get_column_widths()
        for col_idx, (col_letter, width) in enumerate(zip('ABCDEFGHIJKLMNOPQRSTUVWXYZ', column_widths.values()), 1):
            if col_idx <= len(column_widths):
                worksheet.column_dimensions[col_letter].width = width
        
        # 格式化标题行
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col)
            self._apply_style(cell, self.formatter.styles['header'])
        
        # 格式化数据行
        for row in range(2, data_rows + 2):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                
                # 根据列类型应用不同格式
                if col == 1:  # 股票代码列
                    self._apply_style(cell, self.formatter.styles['stock_code'])
                elif worksheet.cell(row=1, column=col).value == '最终排名':
                    # 排名列特殊处理
                    if cell.value and int(cell.value) <= 10:
                        self._apply_style(cell, self.formatter.styles['top_rank'])
                    else:
                        self._apply_style(cell, self.formatter.styles['data'])
                elif isinstance(cell.value, (int, float)):
                    self._apply_style(cell, self.formatter.styles['number'])
                else:
                    self._apply_style(cell, self.formatter.styles['data'])
        
        # 设置冻结窗格
        freeze_panes = self.output_config['excel_settings'].get('freeze_panes')
        if freeze_panes:
            worksheet.freeze_panes = worksheet.cell(row=freeze_panes[0] + 1, column=freeze_panes[1] + 1)
        
        # 启用自动筛选
        if self.output_config['excel_settings'].get('auto_filter', False):
            worksheet.auto_filter.ref = f"A1:{worksheet.cell(row=1, column=worksheet.max_column).coordinate}"
        
        # 添加统计信息工作表
        self._add_statistics_sheet(worksheet.parent, statistics)
    
    def _apply_style(self, cell, style_dict: Dict[str, Any]):
        """应用单元格样式"""
        if not EXCEL_AVAILABLE:
            return
        
        for style_type, style_value in style_dict.items():
            setattr(cell, style_type, style_value)
    
    def _add_statistics_sheet(self, workbook, statistics: Dict[str, Any]):
        """添加统计信息工作表"""
        if not EXCEL_AVAILABLE or not statistics:
            return
        
        try:
            stats_ws = workbook.create_sheet("统计信息")
            
            # 写入基本统计
            stats_ws.cell(row=1, column=1, value="统计项目")
            stats_ws.cell(row=1, column=2, value="数值")
            
            row = 2
            basic_stats = [
                ("计算日期", statistics.get("calculation_date", "")),
                ("总股票数", statistics.get("total_stocks", 0)),
                ("大盘股数量", statistics.get("大盘股数量", 0)),
                ("小盘股数量", statistics.get("小盘股数量", 0)),
            ]
            
            for stat_name, stat_value in basic_stats:
                stats_ws.cell(row=row, column=1, value=stat_name)
                stats_ws.cell(row=row, column=2, value=stat_value)
                row += 1
            
            # 写入得分统计
            if "得分统计" in statistics:
                row += 1
                stats_ws.cell(row=row, column=1, value="得分统计")
                row += 1
                
                score_stats = statistics["得分统计"]
                for stat_name, stat_value in score_stats.items():
                    stats_ws.cell(row=row, column=1, value=stat_name)
                    stats_ws.cell(row=row, column=2, value=f"{stat_value:.4f}" if isinstance(stat_value, (int, float)) else str(stat_value))
                    row += 1
            
            # 设置列宽
            stats_ws.column_dimensions['A'].width = 20
            stats_ws.column_dimensions['B'].width = 15
            
            # 格式化标题
            for col in [1, 2]:
                cell = stats_ws.cell(row=1, column=col)
                self._apply_style(cell, self.formatter.styles['header'])
                
        except Exception as e:
            logger.warning(f"添加统计信息失败: {e}")

# ==============================================================================
# 3. 主Excel工具类
# ==============================================================================

class ExcelUtils:
    """Excel工具主类"""
    
    def __init__(self):
        self.writer = ExcelWriter()
        self.formatter = ExcelFormatter()
        self.output_config = get_output_config()
    
    def export_scoring_results(self, results: pd.DataFrame,
                             calculation_date: str,
                             statistics: Optional[Dict[str, Any]] = None,
                             output_path: Optional[str] = None) -> Dict[str, Any]:
        """
        导出评分结果到Excel
        
        Args:
            results: 评分结果DataFrame
            calculation_date: 计算日期
            statistics: 统计信息
            output_path: 指定输出路径
            
        Returns:
            Dict: 导出结果信息
        """
        try:
            if results.empty:
                return {
                    "success": False,
                    "error": "评分结果为空"
                }
            
            # 写入Excel文件
            output_file = self.writer.write_scoring_results(
                results=results,
                calculation_date=calculation_date,
                statistics=statistics or {},
                output_path=output_path
            )
            
            # 生成导出摘要
            export_summary = {
                "success": True,
                "output_file": output_file,
                "calculation_date": calculation_date,
                "total_stocks": len(results),
                "file_size": self._get_file_size(output_file),
                "columns_count": len(results.columns),
                "export_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            
            logger.info(f"✅ Excel导出成功: {output_file}")
            return export_summary
            
        except Exception as e:
            error_msg = f"Excel导出失败: {e}"
            logger.error(error_msg)
            return {
                "success": False,
                "error": error_msg
            }
    
    def _get_file_size(self, file_path: str) -> str:
        """获取文件大小"""
        try:
            size_bytes = Path(file_path).stat().st_size
            if size_bytes < 1024:
                return f"{size_bytes} B"
            elif size_bytes < 1024 * 1024:
                return f"{size_bytes / 1024:.1f} KB"
            else:
                return f"{size_bytes / (1024 * 1024):.1f} MB"
        except:
            return "未知"
    
    def create_summary_report(self, batch_results: Dict[str, Any], 
                            output_dir: Optional[str] = None) -> str:
        """
        创建批量计算的汇总报告
        
        Args:
            batch_results: 批量计算结果
            output_dir: 输出目录
            
        Returns:
            str: 报告文件路径
        """
        if output_dir is None:
            output_dir = self.output_config['output_directory']
        
        # 创建汇总DataFrame
        summary_data = []
        
        for date, result in batch_results.get('results', {}).items():
            if result.get('success', False):
                summary_data.append({
                    '计算日期': date,
                    '状态': '成功',
                    '股票数量': result.get('total_calculated', 0),
                    '大盘股': result.get('statistics', {}).get('大盘股数量', 0),
                    '小盘股': result.get('statistics', {}).get('小盘股数量', 0),
                    '平均得分': result.get('statistics', {}).get('得分统计', {}).get('平均分', 0)
                })
            else:
                summary_data.append({
                    '计算日期': date,
                    '状态': '失败',
                    '错误信息': result.get('error', '未知错误')
                })
        
        summary_df = pd.DataFrame(summary_data)
        
        # 生成输出路径
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = Path(output_dir) / f"批量计算汇总_{timestamp}.xlsx"
        
        # 写入Excel
        summary_df.to_excel(output_path, index=False, sheet_name='批量计算汇总')
        
        return str(output_path)
    
    def validate_excel_output(self, file_path: str) -> Dict[str, Any]:
        """验证Excel输出文件"""
        validation_result = {
            "file_exists": False,
            "readable": False,
            "sheets": [],
            "data_rows": 0,
            "columns": [],
            "errors": []
        }
        
        try:
            file_path = Path(file_path)
            validation_result["file_exists"] = file_path.exists()
            
            if not validation_result["file_exists"]:
                validation_result["errors"].append("文件不存在")
                return validation_result
            
            # 尝试读取文件
            df = pd.read_excel(file_path, sheet_name=0)
            validation_result["readable"] = True
            validation_result["data_rows"] = len(df)
            validation_result["columns"] = df.columns.tolist()
            
            # 检查工作表
            if EXCEL_AVAILABLE:
                wb = openpyxl.load_workbook(file_path)
                validation_result["sheets"] = wb.sheetnames
            
        except Exception as e:
            validation_result["errors"].append(f"文件验证失败: {e}")
        
        return validation_result

# ==============================================================================
# 4. 便捷函数
# ==============================================================================

def export_to_excel(results: pd.DataFrame, 
                   calculation_date: str,
                   statistics: Optional[Dict[str, Any]] = None,
                   output_path: Optional[str] = None) -> Dict[str, Any]:
    """
    便捷的Excel导出函数
    
    Args:
        results: 评分结果
        calculation_date: 计算日期
        statistics: 统计信息
        output_path: 输出路径
        
    Returns:
        Dict: 导出结果
    """
    excel_utils = ExcelUtils()
    return excel_utils.export_scoring_results(results, calculation_date, statistics, output_path)

def check_excel_dependencies() -> Dict[str, bool]:
    """检查Excel相关依赖"""
    dependencies = {
        "openpyxl": EXCEL_AVAILABLE,
        "pandas": True,  # 如果能运行到这里，pandas肯定可用
    }
    
    return dependencies

# ==============================================================================
# 5. 使用示例和测试
# ==============================================================================

if __name__ == "__main__":
    logger.info("=== Excel工具测试 ===")
    
    try:
        # 检查依赖
        deps = check_excel_dependencies()
        print("依赖检查:")
        for dep, available in deps.items():
            print(f"  {dep}: {'✅' if available else '❌'}")
        
        # 创建测试数据
        test_data = pd.DataFrame({
            'ts_code': ['000001.SZ', '000002.SZ', '600000.SH'],
            'stock_name': ['平安银行', '万科A', '浦发银行'],
            'industry': ['银行', '房地产', '银行'],
            'ROE': [0.1234, 0.0987, 0.1456],
            'ROA': [0.0567, 0.0432, 0.0678],
            'profitability_score': [1.23, -0.45, 0.89],
            'growth_score': [0.56, 1.23, -0.34],
            'safety_score': [-0.12, 0.67, 0.23],
            'final_score': [0.89, 0.45, 0.23],
            'final_rank': [1, 2, 3],
            'market_cap_group': ['large_cap', 'large_cap', 'large_cap']
        })
        
        test_statistics = {
            "calculation_date": "2022-05-01",
            "total_stocks": 3,
            "大盘股数量": 3,
            "小盘股数量": 0,
            "得分统计": {
                "平均分": 0.52,
                "标准差": 0.33,
                "最高分": 0.89,
                "最低分": 0.23
            }
        }
        
        # 测试Excel导出
        excel_utils = ExcelUtils()
        result = excel_utils.export_scoring_results(
            results=test_data,
            calculation_date="2022-05-01",
            statistics=test_statistics
        )
        
        if result["success"]:
            print(f"\n✅ Excel导出测试成功!")
            print(f"输出文件: {result['output_file']}")
            print(f"文件大小: {result['file_size']}")
            print(f"股票数量: {result['total_stocks']}")
            
            # 验证输出文件
            validation = excel_utils.validate_excel_output(result['output_file'])
            print(f"\n文件验证:")
            print(f"文件存在: {'✅' if validation['file_exists'] else '❌'}")
            print(f"可读取: {'✅' if validation['readable'] else '❌'}")
            print(f"数据行数: {validation['data_rows']}")
            print(f"列数: {len(validation['columns'])}")
            
        else:
            print(f"❌ Excel导出测试失败: {result['error']}")
        
        print("\n=== 测试完成 ===")
        
    except Exception as e:
        logger.error(f"测试失败: {e}")
        raise

# ==============================================================================
# 6. 导出接口
# ==============================================================================

__all__ = [
    'ExcelUtils',
    'ExcelFormatter',
    'ExcelWriter',
    'export_to_excel',
    'check_excel_dependencies'
]